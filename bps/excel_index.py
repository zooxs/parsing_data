from openpyxl import load_workbook

# open & select the worksheet
wb = load_workbook('Produksi Palawija.xlsx', data_only=True)
sheets_range = wb.sheetnames
ws = wb.active

def letter2number(value):
    letter = 'a,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v,w,x,y,z'
    code = letter.upper().split(',')
    n = len(code)
    
    if type(value) == str:
        return code.index(value.upper()) + 1
    elif type(value) == int:
        return code[value-1]

def ColumnVal(ls):
    values = [letter2number(i.split(':')[1][0]) for i in ls]
    return values

# list for merged cell in worksheet
ls_ranges = [str(i) for i in ws.merged_cells.ranges]

# function to determine the sector in the whorksheet
def sector(ls):
    """
    sector A is cell that contain in column A,
    this sector has max number of row and below 
    constrain for data that could be parsed.
    sector B is cell that contain in column B,
    this sector has several value like title of table,
    classification for data and contain data itself.
    """
    sector_a = [a for a in ls if 'A' in a]
    sector_b = []
    for b in ls:
        b1, b2 = b.split(':')
        if (len(b1) == 2) & (b1[-1] == b2[-1]):
            sector_b.append(b)

    # for sector A
    row_val = sorted([int(an.split(':')[1][1::]) for an in sector_a ])
    a_val = {"start_row": row_val[0] + 1, "end_row":row_val[1] - 1}

    # for sector B
    b_val = {"cell_title": 'B1'}
    for bn in sector_b:
        bn1, bn2 = bn.split(':')
        if '2' in bn1:
            b_val[f"{bn1}_sub_class"] = [bn1, bn2]
    # sectors[f"{data['a1'].value}"] = [
    #     cell.value for row in data.iter_cols(
    #         min_row=start_a, 
    #         max_row=end_a, 
    #         min_col=1,
    #         max_col=1) 
    #     for cell in row]
    # return [data[key.split(':')[0]].value for key in sector_b]
    a_val.update(b_val)
    return a_val
print(sector(ls_ranges))
